from openpyxl import load_workbook
from datetime import datetime

rut=r'hojaDatosProductos.xlsx'

#funciones alejandra

def consultar_producto(rut:str,extraer:str):
    Archivo_Exccel=load_workbook(rut)
    Hoja_datos=Archivo_Exccel["productos"]
    Hoja_datos=Hoja_datos["A2":"E"+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{"nombre":i[1].value, "categoria":i[2].value, "precio":i[3].value, "cantidad":i[4].value})

    if not(extraer=="todo"):
        info=filtrar(info, extraer)

    for i in info:
        print("****Productos****")
        print("id: "+ str(i)+"\n"+"Nombre: "+str(info[i]["nombre"])+"\n"+"Categoria: "+str(info[i]["categoria"])+"\n"+"Precio: "+str(info[i]["precio"])+"\n"+"Cantidad: "+str(info[i]["cantidad"]))
        print()

    return

def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]["categoria"]==filtro:
            aux.setdefault(i,info[i])
    return aux

def agregar_producto(rut:int,datos:dict):
    Archivo_Exccel=load_workbook(rut)
    Hoja_datos=Archivo_Exccel["productos"]
    Hoja_datos=Hoja_datos["A2":"E"+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Exccel.active

    nombre=2
    categoria=3
    precio=4
    cantidad=5
    for i in Hoja_datos:

        if not(isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=nombre).value=datos["nombre"]
            hoja.cell(row=identificador, column=categoria).value=datos["categoria"]
            hoja.cell(row=identificador, column=precio).value=datos["precio"]
            hoja.cell(row=identificador, column=cantidad).value=datos["cantidad"] 
            break
    Archivo_Exccel.save(rut)
    return

#funciones yoiner

def borrar(ruta,identificador):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['productos']
  hojaDatos=hojaDatos['A2':'E'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  categoria=3
  precio=4
  cantidad=5
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      hoja.cell(row=fila,column=1).value=''
      hoja.cell(row=fila,column=nombre).value=''
      hoja.cell(row=fila,column=categoria).value=''
      hoja.cell(row=fila,column=precio).value=''
      hoja.cell(row=fila,column=cantidad).value=''
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

def actualizar(ruta:str,identificador:int,datosActualizados:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['productos']
  hojaDatos=hojaDatos['A2':'E'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  categoria=3
  precio=4
  cantidad=5
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      for d in datosActualizados:
        if d=='nombre' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=nombre).value=datosActualizados[d]
        elif d=='categoria' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=categoria).value=datosActualizados[d]
        elif d=='precio' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=precio).value=datosActualizados[d]
        elif d=='cantidad' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=cantidad).value=datosActualizados[d]
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

#While - yoiner
while True:
  print('******************************************')
  print('Indique la accion que desea realizar: \nConsultar: 1\nActualizar: 2\nCrear nuevo producto: 3\nBorrar: 4')
  accion =int(input('Escriba la opcion: '))
  if accion<1 or accion>4:
    print('Comando invalido, por favor eliga una opcion valida')
  elif accion==1:
    opcConsulta=''
    print('Indique la categoria del producto que desea consultar:\nTodos los productos: 1\nComputación: 2\nAlimentario: 3\nHigiene: 4\nEscolar: 5')
    opcConsulta=input('Escriba la categoria que desee consultar: ')
    if opcConsulta=='1':
      print('\n\n** Consultado todos los productos **')
      consultar_producto(rut,'todo')
    elif opcConsulta=='2':
      print('\n\n** Consultado todos los productos **')
      consultar_producto(rut,'computacion')
    elif opcConsulta=='3':
      print('\n\n** Consultado todos los productos **')
      consultar_producto(rut,'alimentario')
    elif opcConsulta=='4':
      print('\n\n** Consultado todos los productos **')
      consultar_producto(rut,'higiene')
    elif opcConsulta=='5':
      print('\n\n** Consultado todos los productos **')
      consultar_producto(rut,'escolar')
  elif accion==2:
    datosActualizados={'nombre':'','categoria':'','precio':'','cantidad':''}
    print('** Actualizar Tarea **\n')
    idActualizar=int(input('Indique el ID de el producto que desea actualizar: '))

    print('\n** Nuevo nombre **\n** Nota: si no desea actualizar el nombre solo oprima ENTER **')
    datosActualizados['nombre']=input('Indique el nuevo nombre de el producto: ')

    print('\n** Nueva categoria **\nComputación: 1\nAlimentario: 2\nHigiene: 3\nEscolar: 4\n** Nota: si no desea actualizar la categoria solo oprima ENTER **')
    
    estadoNuevo=input('Indique el nuevo estado de el producto: ')
    if estadoNuevo=='1':
      datosActualizados['categoria']='computacion'
    elif estadoNuevo=='2':
      datosActualizados['categoria']='alimentario'
    elif estadoNuevo=='3':
      datosActualizados['categoria']='higiene'
    elif estadoNuevo=='4':
      datosActualizados['categoria']='escolar'
    print('\n** Nuevo precio **\n** Nota: si no desea actualizar el precio solo oprima ENTER **')
    datosActualizados['precio']=input('Indique el nuevo precio de el producto: ')

    print('\n** Nueva cantidad **\n** Nota: si no desea actualizar la cantidad solo oprima ENTER **')
    datosActualizados['cantidad']=input('Indique la nueva cantidad de el producto: ')

    actualizar(rut,idActualizar, datosActualizados)
    print()
  elif accion==3:
    datosActualizados={'nombre':'','categoria':'','precio':'','cantidad':''}
    print('** Crear nuevo producto **\n')
    print('** Nombre **\n')
    datosActualizados['nombre']=input('Indique el nombre de el producto: ')

    print('\n** Categoria **\nComputación: 1\nAlimentario: 2\nHigiene: 3\nEscolar: 4')
    
    estadoNuevo=input('Indique la categoria de el producto: ')
    if estadoNuevo=='1':
      datosActualizados['categoria']='computacion'
    elif estadoNuevo=='2':
      datosActualizados['categoria']='alimentario'
    elif estadoNuevo=='3':
      datosActualizados['categoria']='higiene'
    elif estadoNuevo=='4':
      datosActualizados['categoria']='escolar'
    print('\n** Precio **')
    datosActualizados['precio']=input('Indique el precio de el producto: ')
    print('\n** cantidad **')
    datosActualizados['cantidad']=input('Indique el nombre de el producto: ')
    agregar_producto(rut,datosActualizados)
  elif accion==4:
    print('\n** Eliminar producto **')
    iden=int(input('Indique el ID de el producto que desea eliminar: '))
    borrar(rut,iden)